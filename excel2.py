# -------------------------------------- Exercises --------------------------------------
# 2) Read several specific cells from an existing Excel file and store them on variables;

# Imports
import openpyxl

# Creating the Workbook and Worksheet Objects from an existing file
wb = openpyxl.load_workbook(".\\examplesheet\\example.xlsx")
ws = wb.active

# Printing a specific cell
my_cell = ws.cell(row=1, column=2)
print(my_cell.value)

# Defining number of rows/coluns
sheet_rows = ws.max_row
sheet_cols = ws.max_column
print(f"A planilha carregada possui {sheet_rows} linhas e {sheet_cols} colunas! Ciência...\n")


# Printing all columns' names
def get_columns(mystring=""):

    for i in range(1, sheet_cols + 1):
        current_cell = ws.cell(row=1, column=i)
        print(current_cell.value)

        if current_cell.value == mystring:
            return current_cell.value       # Return something useful here to build a string with desired data


print("Colunas da planilha:\n")
get_columns()





