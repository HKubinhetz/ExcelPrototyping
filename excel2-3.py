# -------------------------------------- Exercises --------------------------------------
# 2) Read several specific cells from an existing Excel file and store them on variables;
# With openpyxl

# Imports
import openpyxl

# Creating the Workbook and Worksheet Objects from an existing file
wb = openpyxl.load_workbook(".\\examplesheet\\example.xlsx")
ws = wb.active

# Printing a specific cell
my_cell = ws.cell(row=1, column=2)
print(my_cell.value)

# Defining number of rows/coluns
sheet_rows = ws.max_row
sheet_cols = ws.max_column
print(f"A planilha carregada possui {sheet_rows} linhas e {sheet_cols} colunas! Ciência...\n")


# Printing all columns' names
def get_columns(mystring=""):

    for i in range(1, sheet_cols + 1):
        current_cell = ws.cell(row=1, column=i)
        print(current_cell.value)

        if current_cell.value == mystring:
            return i       # Return something useful here to build a string with desired data


# Grabbing relevant indexes
cdie_index = get_columns("INSTALAÇÃO")
name_index = get_columns("NOME/SOBRENOME")
tech_index = get_columns("Tecnologia")

# Listing a readable relation
print("Lista de Clientes")

for i in range(1, sheet_cols-1):
    current_cdie = ws.cell(row=i, column=cdie_index).value
    current_name = ws.cell(row=i, column=name_index).value
    current_tech = ws.cell(row=i, column=tech_index).value
    print(f"{current_cdie} - {current_name}: Possui tecnologia {current_tech}.")



